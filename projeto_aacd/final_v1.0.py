import xlsxwriter
import openpyxl
import os

print("1 - Cadastrar nova empresa")
print("2 - Cadastrar novo cofre")
print("3 - Recolhimento do cofre")
print("4 - Cadastrar valor recolhido")
print("5 - Editar Informações da empresa")
print("6 - Sair")

perguntaI = input("\nO que deseja fazer? ")

def conf(arquivo):    

    workbook = xlsxwriter.Workbook(arquivo)
    worksheet = workbook.add_worksheet()
    cell_format = workbook.add_format({'bold': True})
    red = workbook.add_format({'font_color':'red'})
    green = workbook.add_format({'font_color':'green'})

    for r in range(20):
        worksheet.set_column(0,1,20)
        worksheet.set_column(2,3,10)
        worksheet.set_column(4,5,20)
        worksheet.set_column(6,7,18)
        worksheet.set_row(0, 30, cell_format)
        worksheet.set_row(1, 30, cell_format)
        worksheet.set_row(2, 18)
        worksheet.set_row(3, 18, green)
        worksheet.set_row(r+3, 18, red)

    ## Entrada de dados ##
        
    endereco=input('Digite o endereço da empresa: ')
    telefone=input('Digite o telefone da empresa: ')
    
    ## Cabeçalho ##
    worksheet.write(0,0,nomeArquivo)
    worksheet.write(0,1,endereco)
    worksheet.write(1,0,telefone)
    worksheet.write(1,1,'Previsão')

    ## Fim de criação de Arquivo ##

    workbook.close()
    
    ## Configurações da Planiha de empresas ##
    try:
        os.chdir('C:\\Users\\Aluno\\Desktop\\Prototipo')
        wb2=openpyxl.load_workbook('Empresas.xlsx')
        sheet2 = wb2['Sheet1']    
    except FileNotFoundError:
        workbook2 = xlsxwriter.Workbook('Empresas.xlsx')
        worksheet2 = workbook2.add_worksheet()
        worksheet2.set_column('A:A',20)
        worksheet2.set_column('B:B',25)
        worksheet2.set_column('C:C',12)
        worksheet2.set_column('D:D',10)
        worksheet2.set_column('E:E',15)
        worksheet2.write('A1', "Nome da Empresa")
        worksheet2.write('B1', "Endereço")
        worksheet2.write('C1', "Telefone")
        worksheet2.write('D1', "Cofre Atual")
        worksheet2.write('E1', "Previsão Recol.")
        
        workbook2.close()

    os.chdir('C:\\Users\\Aluno\\Desktop\\Prototipo')
    wb2=openpyxl.load_workbook('Empresas.xlsx')
    sheet2 = wb2['Sheet1']
    
    ## Inicio da planilha Empresas ##

    for l in range(10):
            if sheet2.cell(11-l,2).value!=None:
                for i in range (1,6):
                    sheet2.cell(12-l,i).value=sheet2.cell(11-l,i).value 
    sheet2.cell(2,1).value= nomeArquivo
    sheet2.cell(2,2).value= endereco
    sheet2.cell(2,3).value= telefone
    
    wb2.save('Empresas.xlsx')    

    print('Fim da configuração do Arquivo.')

def cadastroCofre(arquivo):
    os.chdir('C:\\Users\\Aluno\\Desktop\\Prototipo')
    wb=openpyxl.load_workbook(arquivo)
    sheet = wb['Sheet1']

    sheet.cell(3,1).value='Status do Cofre'
    sheet.cell(3,2).value='N° do cofre'
    sheet.cell(3,3).value='Instalação'
    sheet.cell(3,4).value='Recolhimento'
    sheet.cell(3,5).value='Voluntario Inst.'
    sheet.cell(3,6).value='Voluntario Reco.'
    sheet.cell(3,7).value='Valor Recolhido'     

            ## Dados do cofre Novo ##

    cofreAtual=input('Digite o número do cofre atual: ')
    dataInst=input('Digite a data de instalção: ')
    nomeVolun=input('Digite o nome do voluntario que instalou o cofre: ')
    statusCA='ATIVO'
            
        ## Pulando linhas ##
    for l in range(10):
        if sheet.cell(14-l,2).value!=None:
            for i in range (1,6):
                sheet.cell(15-l,i).value=sheet.cell(14-l,i).value 
   
    for i in range(1,6):
        sheet.cell(5,i+1).value=sheet.cell(4,i+1).value
               
    sheet.cell(4,1).value=statusCA
    sheet.cell(4,2).value=cofreAtual
    sheet.cell(4,3).value=dataInst
    sheet.cell(4,4).value=None
    sheet.cell(4,5).value=nomeVolun
    sheet.cell(4,6).value=None
    sheet.cell(4,7).value=None
    wb.save(arquivo)
        

while perguntaI != '6':
    if perguntaI == "1" or perguntaI == "Cadastrar nova empresa":
        nomeArquivo = input("Digite o nome do arquivo: ")
        arquivo = nomeArquivo+'.xlsx'
        conf(arquivo)

    elif perguntaI == '2' or perguntaI == 'Cadastrar novo cofre':
        nomeArquivo = input("Nome da empresa: ")
        arquivo = nomeArquivo + '.xlsx'
        try:
            os.chdir('C:\\Users\\Aluno\\Desktop\\Prototipo')
            wb=openpyxl.load_workbook(arquivo)
            sheet = wb['Sheet1']
        except FileNotFoundError:
            perguntaC = input("Empresa não cadastrada, deseja cadastrar? ")
            if perguntaC == 'Sim' or perguntaC == 'sim':
                conf(arquivo)
            else:
                break
        cadastroCofre(arquivo)

        os.chdir('C:\\Users\\Aluno\\Desktop\\Prototipo')
        wb2=openpyxl.load_workbook('Empresas.xlsx')
        sheet2 = wb2['Sheet1']  
        for y in range(1,11):
            if sheet2.cell(y, 1).value == nomeArquivo:
                sheet2.cell(y, 4).value = cofreAtual
                           
        wb2.save('Empresas.xlsx')

    elif perguntaI == '3' or perguntaI == 'Recolhimento do cofre':
        nomeArquivo = input('Nome da Empresa: ')
        arquivo = nomeArquivo + '.xlsx'
        try:
            os.chdir('C:\\Users\\Aluno\\Desktop\\Prototipo')
            wb=openpyxl.load_workbook(arquivo)
            sheet = wb['Sheet1']
        except FileNotFoundError:
            perguntaD = input('Empresa não cadastrada, deseja cadastrar a empresa?')
            if perguntaD == 'Sim' or perguntaD == 'sim':
                conf(arquivo)
            else:
                break
            
        cofreAtual=input('Digite o número do cofre recolhido: ')
        dataRecol=input('Digite a data de recolhimento: ')
        nomeVolunR=input('Digite o nome do voluntario que recolheu o cofre: ')
        statusCR='RECOLHIDO'
            
        os.chdir('C:\\Users\\Aluno\\Desktop\\Prototipo')
        wb=openpyxl.load_workbook(arquivo)
        sheet = wb['Sheet1']
        y=1
        for x in range(1,40):
            if sheet.cell(x, 2).value==cofreAtual:
                sheet.cell(x,1).value=statusCR
                sheet.cell(x,4).value=dataRecol
                sheet.cell(x,6).value=nomeVolunR
                y=0

        if y==1:    
            preguntaRecol = input("Cofre não foi cadastrado, deseja cadastrar?")
            if preguntaRecol == 'Sim' or preguntaRecol == 'sim':
              cadastroCofre(arquivo)
            else:
                break  

        wb.save(arquivo)
    
    elif perguntaI == '4' or perguntaI == 'Cadastrar valor recolhido':
        nomeArquivo = input('Nome da Empresa: ')
        arquivo = nomeArquivo + '.xlsx'
        cofreV = input('Qual o numero do cofre?')
        valorRecolhido = input('Qual o valor Recolhiddo?')

        os.chdir('C:\\Users\\Aluno\\Desktop\\Prototipo')
        wb=openpyxl.load_workbook(arquivo)
        sheet = wb['Sheet1']

        for x in range(1,40):
            if sheet.cell(x, 2).value == cofreV :
                sheet.cell(x,7).value = valorRecolhido
        
        wb.save(arquivo)
    
    elif perguntaI == '5' or perguntaI == 'Editar Informações da empresa':
        nomeArquivo = input('Nome da Empresa: ')
        arquivo = nomeArquivo + '.xlsx'
        
        os.chdir('C:\\Users\\Aluno\\Desktop\\Prototipo')
        wb=openpyxl.load_workbook(arquivo)
        sheet = wb['Sheet1']

        perguntaEndereco = input('Deseja mudar o endereço da empresa? ')
        if perguntaEndereco == 'Sim' or perguntaEndereco == 'sim':   
            endereco=input('Digite o endereço da empresa: ')
            sheet.cell(1,2).value = None
            sheet.cell(1,2).value = endereco
        else:
            print('ok')
        
        perguntaTelefone = input('Deseja mudar o telefone da empresa? ')
        if perguntaTelefone == 'Sim' or perguntaTelefone == 'sim':
            telefone=input('Digite o telefone da empresa: ')
            sheet.cell(2,1).value = None
            sheet.cell(2,1).value = telefone
        else:
            print('ok')
        
        wb.save(arquivo)
        
        print('Informações de ', nomeArquivo, ' foram atualizadas.')
       
    elif perguntaI == '6' or perguntaI == 'Sair':
        break
    
    print("\n\n1 - Cadastrar nova empresa")
    print("2 - Cadastrar novo cofre")
    print("3 - Recolhimento do cofre")
    print("4 - Cadastrar valor recolhido")
    print("5 - Editar Informações da empresa")
    print("6 - Sair")
    perguntaI = input("\nO que deseja fazer? ")
    
